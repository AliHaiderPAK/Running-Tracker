from tkinter import *
from tkinter.font import ITALIC
from turtle import color
from tkcalendar import Calendar
from datetime import date
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg)
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt1
from matplotlib import style
from statistics import mean
from tkinter import ttk

style.use('ggplot')
cellc=1
def toexcel():
    wb=load_workbook('data.xlsx')
    ws=wb.active
    ws.append([cal.get_date(),distance.get(),hourentry.get()+":"+minentry.get()+":"+secentry.get()])
    wb.save("data.xlsx")
    update()
def datesfromexcel():
    wb=load_workbook('data.xlsx')
    ws=wb.active
    i=1
    dates=[]
    while (ws["A"+str(i)].value!=None):
        dates.append(ws["A"+str(i)].value)
        i+=1
    wb.save("data.xlsx")
    return dates
def distsfromexcel():
    wb=load_workbook('data.xlsx')
    ws=wb.active
    i=1
    dists=[]
    while (ws["A"+str(i)].value!=None):
        
        dists.append(float(ws["B"+str(i)].value))
        i+=1
    wb.save("data.xlsx")
    return dists
def timesfromexcel():
    wb=load_workbook('data.xlsx')
    ws=wb.active
    i=1
    times=[]
    while (ws["A"+str(i)].value!=None):
        h, m, s =ws["C"+str(i)].value.split(':')
        minutes=(int(h) * 3600 + int(m) * 60 + int(s))/60
        times.append(minutes)
        i+=1
    wb.save("data.xlsx")
    return times
def distancegraph(x,y):
    fig = Figure(figsize = (5.3,4.1),
                 dpi = 100)
    plt1 = fig.add_subplot(111)
    plt1.grid(True, color='k')
    plt1.scatter(x,y, color="blue")
    plt1.plot(x,y,linewidth=2, color="blue")
    canvas = FigureCanvasTkAgg(fig,
                               master = window)
    canvas.draw()
    canvas.get_tk_widget().place(relx=0.5, rely=0.5, anchor=CENTER)
    canvas.get_tk_widget().place(relx=0.1649, rely=0.701)
    candist=Label(window,
                text="Distance(kms)",
                font=("Arial", 15))
    candist.place(x=200,y=425)
def timegraph(x,y):
    fig = Figure(figsize = (5.3,4.1),
                 dpi = 100)
    plt1 = fig.add_subplot(111)
    plt1.grid(True, color='k')
    plt1.scatter(x,y, color="green")
    plt1.plot(x,y,linewidth=2, color="green")
    canvas = FigureCanvasTkAgg(fig,
                               master = window)  
    canvas.draw()
    canvas.get_tk_widget().place(relx=0.498, rely=0.701, anchor=CENTER)
    cantime=Label(window,
                text="Time(mins)",
                font=("Arial", 15))
    cantime.place(x=760,y=425)

def speedgraph(x,y):
    fig = Figure(figsize = (5.3,4.1),
                 dpi = 100)
    plt1 = fig.add_subplot(111)
    plt1.grid(True, color='k')
    plt1.scatter(x,y)
    plt1.plot(x,y,linewidth=2)
    canvas = FigureCanvasTkAgg(fig,
                               master = window)  
    canvas.draw()
    canvas.get_tk_widget().place(relx=0.831, rely=0.701, anchor=CENTER)
    canspeed=Label(window,
                text="Speed(kmph)",
                font=("Arial", 15))
    canspeed.place(x=1270,y=425)
def dellast():
    wb=load_workbook('data.xlsx')
    ws=wb.active
    i=1
    while (ws["A"+str(i)].value!=None):
         i+=1
    ws["A"+str(i-1)].value=None
    ws["B"+str(i-1)].value=None
    ws["C"+str(i-1)].value=None
    wb.save("data.xlsx")
    update()
def update():
    try:
        distlist=(distsfromexcel())
        datelist=(datesfromexcel())
        timelist=(timesfromexcel())
        gdistlist=distlist[-7:]
        gdatelist=datelist[-7:]
        gtimelist=timelist[-7:]
        speedlist=[]
        for i in range(len(distlist)):
            distlist[i]=float(distlist[i])
        for i in range(len(distlist)):    
            speedlist.append((distlist[i]/timelist[i])*60)
        gspeedlist=speedlist[-7:]
        distancegraph(gdatelist,gdistlist)
        timegraph(gdatelist,gtimelist)
        speedgraph(gdatelist,gspeedlist)
        runslabel.config(text=len(datelist))
        mindistlabel.config(text=min(distlist))
        maxdistlabel.config(text=max(distlist))
        avgdistlabel.config(text=round(mean(distlist),2))
        distlabel.config(text=round(sum(distlist),2))
        minspeedlabel.config(text=round(min(speedlist),2))
        maxspeedlabel.config(text=round(max(speedlist),2))
        avgspeedlabel.config(text=round(mean(speedlist),2))
        timelabel.config(text=round(sum(timelist),2))
        mintime.config(text=min(timelist))
        maxtime.config(text=round(max(timelist),2))
        avgtime.config(text=round(mean(timelist),2))    
    except ValueError:
        distlist=(distsfromexcel())
        datelist=(datesfromexcel())
        timelist=(timesfromexcel())
        speedlist=[]
        for i in range(len(distlist)):
            distlist[i]=float(distlist[i])
        for i in range(len(distlist)):    
            speedlist.append((distlist[i]/timelist[i])*60)
        distancegraph(datelist,distlist)
        timegraph(datelist,timelist)
        speedgraph(datelist,speedlist)
        runslabel.config(text="0")
        mindistlabel.config(text="0")
        maxdistlabel.config(text="0")
        avgdistlabel.config(text="0")
        distlabel.config(text="0")
        minspeedlabel.config(text="0")
        maxspeedlabel.config(text="0")
        avgspeedlabel.config(text="0")
        timelabel.config(text="0")
        mintime.config(text="0")
        maxtime.config(text="0")
        avgtime.config(text="0")
def fullgraph():
    def distancegraphf(x,y):
        fig = Figure(figsize = (19,9),
                    dpi = 100)
        plt1 = fig.add_subplot(111)
        plt1.grid(True, color='k')
        plt1.scatter(x,y, color="blue")
        plt1.plot(x,y,linewidth=2, color="blue")
        canvas = FigureCanvasTkAgg(fig,
                                master = frame1)
        canvas.draw()
        canvas.get_tk_widget().place(relx=0.5, rely=0.5, anchor=CENTER)
        canvas.get_tk_widget().place(relx=0.5, rely=0.48)
        candist=Label(frame1,
                    text="Distance(kms)",
                    font=("Arial", 15))
        candist.place(x=700,y=20)
    def timegraphf(x,y):
        fig = Figure(figsize = (19,9),
                    dpi = 100)
        plt1 = fig.add_subplot(111)
        plt1.grid(True, color='k')
        plt1.scatter(x,y, color="green")
        plt1.plot(x,y,linewidth=2, color="green")
        canvas = FigureCanvasTkAgg(fig,
                                master = frame2)  
        canvas.draw()
        canvas.get_tk_widget().place(relx=0.5, rely=0.48, anchor=CENTER)
        cantime=Label(frame2,
                    text="Time(mins)",
                    font=("Arial", 15))
        cantime.place(x=700,y=20)

    def speedgraphf(x,y):
        fig = Figure(figsize = (19,9),
                    dpi = 100)
        plt1 = fig.add_subplot(111)
        plt1.grid(True, color='k')
        plt1.scatter(x,y)
        plt1.plot(x,y,linewidth=2)
        canvas = FigureCanvasTkAgg(fig,
                                master = frame3)  
        canvas.draw()
        canvas.get_tk_widget().place(relx=0.5, rely=0.48, anchor=CENTER)
        canspeed=Label(frame3,
                    text="Speed(kmph)",
                    font=("Arial", 15))
        canspeed.place(x=700,y=20)
    fulgra=Tk()
    w= window.winfo_screenwidth() 
    h= window.winfo_screenheight()
    fulgra.geometry("%dx%d" % (w, h))
    fulgra.title("Full Graphs")
    notebook=ttk.Notebook(fulgra)
    frame1 = ttk.Frame(notebook, width=w, height=h)
    frame2 = ttk.Frame(notebook, width=w, height=h)
    frame3 = ttk.Frame(notebook, width=w, height=h)
    distlist=(distsfromexcel())
    datelist=(datesfromexcel())
    timelist=(timesfromexcel())
    speedlist=[]
    for i in range(len(distlist)):
        distlist[i]=float(distlist[i])
    for i in range(len(distlist)):    
        speedlist.append((distlist[i]/timelist[i])*60)
    distancegraphf(datelist,distlist)
    timegraphf(datelist,timelist)
    speedgraphf(datelist,speedlist)
    frame1.pack(fill='both', expand=True)
    frame2.pack(fill='both', expand=True)
    frame3.pack(fill='both', expand=True)
    notebook.add(frame1, text='Distance')
    notebook.add(frame2, text='Time')
    notebook.add(frame3, text='Speed')
    notebook.pack()
    fulgra.mainloop()
today = date.today()
today=str(today)
y=today[0:4:1]
m=today[5:7:1]
d=today[8:10:1]
y=int(y)
m=int(m)
d=int(d)
window=Tk()
width= window.winfo_screenwidth() 
height= window.winfo_screenheight()
window.geometry("%dx%d" % (width, height))
window.title("Running Tracker")
window.config(background="#c7fffd")
label=Label(window,
            text="R U N N I N G   T R A C K E R",
            font=("ALGERIAN",70,ITALIC),
            width=21,
            height=1,
            bg="black",
            fg="#5eff00")
label.pack()
dataentry=Label(window,
                width=75,
                height=20,
                bd=1,
                relief=RAISED,
                bg="#9eff99")
dataentry.place(x=0,y=110)
headlabel=Label(dataentry,
                text="New Data",
                font=("Arial", 21),
                bd=2,
                relief=RAISED,
                bg="#51ff47")
headlabel.place(x=180,y=2)
distlabel=Label(dataentry,
                text="Distance:",
                font=("Arial", 15),
                bg="#9eff99")
distlabel.place(x=15,y=90)
distance=Entry(dataentry,
                font=("Digital-7", 18),
                width=4,
                fg="#00ff22",
                bg="black")
distance.place(x=105,y=90)
distance.insert(END,"0")
kmlabel=Label(dataentry,
                text="km",
                font=("Arial", 12),
                bg="#9eff99")
kmlabel.place(x=115,y=65)
timelabel=Label(dataentry,
                text="Time:",
                font=("Arial", 15),
                bg="#9eff99")
timelabel.place(x=15,y=200)
hourentry=Entry(dataentry,
                font=("Digital-7", 18),
                width=2,
                fg="#00ff22",
                bg="black")
hourentry.place(x=70,y=200)
hourentry.insert(END,"0")
minentry=Entry(dataentry,
                font=("Digital-7", 18),
                width=2,
                bg="black",
                fg="#00ff22")
minentry.place(x=100,y=200)
minentry.insert(END,"0")
secentry=Entry(dataentry,
                font=("Digital-7", 18),
                width=2,
                bg="black",
                fg="#00ff22",)
secentry.place(x=130,y=200)
secentry.insert(END,"0")
hmslabel=Label(dataentry,
                text="h     m     s",
                font=("Arial", 12),
                bg="#9eff99")
hmslabel.place(x=75,y=175)
datelabel=Label(dataentry,
                text="Date:",
                font=("Arial", 15),
                bg="#9eff99")
datelabel.place(x=200,y=60)
cal = Calendar(dataentry,
                selectmode = 'day',
			    year = y,
                month = m,
			    day = d)
cal.place(x=260,y=65)
submitbut=Button(dataentry,
                 text="Submit",
                 font=("Arial", 15),
                 bg="#808080",
                 pady=2,
                 padx=2,
                 command=toexcel)
submitbut.place(x=190,y=257)
dellastent=Button(dataentry,
                  text="Delete Last Entry",
                  font=("Arial", 15),
                  bg="#808080",
                  pady=2,
                  padx=2,
                  command=dellast)
dellastent.place(x=290, y=257)
stats=Label(window,
            width=152,
            height=20,
            bd=1,
            relief=RAISED,
            bg="#00ff95")
stats.place(x=531,y=110)
statslabel=Label(stats,
                text="Statistics",
                font=("Arial", 21),
                bd=2,
                relief=RAISED,
                bg="#51ff47")
statslabel.place(x=470,y=2)
trunslabel=Label(stats,
                text="Total Runs:",
                font=("Arial", 15),
                bg="#00ff95")
trunslabel.place(x=5,y=80)
runslabel=Label(stats,
                text="0",
                font=("Digital-7", 20),
                bg="black",
                fg="#00ff22",
                width=6,
                height=1)
runslabel.place(x=155,y=80)
tdistlabel=Label(stats,
                text="Total Distance:",
                font=("Arial", 15),
                bg="#00ff95")
tdistlabel.place(x=5,y=130)
distlabel=Label(stats,
                text="0",
                font=("Digital-7", 20),
                bg="black",
                fg="#00ff22",
                width=6,
                height=1)
distlabel.place(x=155,y=130)
ttimelabel=Label(stats,
                text="Total Time:",
                font=("Arial", 15),
                bg="#00ff95")
ttimelabel.place(x=5,y=180)
timelabel=Label(stats,
                text="0",
                font=("Digital-7", 20),
                bg="black",
                fg="#00ff22",
                width=6,
                height=1)
timelabel.place(x=155,y=180)
mintimelabel=Label(stats,
                text="Minimum Time:",
                font=("Arial", 15),
                bg="#00ff95")
mintimelabel.place(x=550,y=80)
mintime=Label(stats,
                text="0",
                font=("Digital-7", 20),
                bg="black",
                fg="#00ff22",
                width=4,
                height=1)
mintime.place(x=700,y=80)
maxtimelabel=Label(stats,
                text="Maximum Time:",
                font=("Arial", 15),
                bg="#00ff95")
maxtimelabel.place(x=550,y=130)
maxtime=Label(stats,
                text="0",
                font=("Digital-7", 20),
                bg="black",
                fg="#00ff22",
                width=4,
                height=1)
maxtime.place(x=700,y=130)
avgtimelabel=Label(stats,
                text="Average Time:",
                font=("Arial", 15),
                bg="#00ff95")
avgtimelabel.place(x=550,y=180)
avgtime=Label(stats,
                text="0",
                font=("Digital-7", 20),
                bg="black",
                fg="#00ff22",
                width=4,
                height=1)
avgtime.place(x=700,y=180)
mindist=Label(stats,
                text="Minimum Distance:",
                font=("Arial", 15),
                bg="#00ff95")
mindist.place(x=280,y=80)
mindistlabel=Label(stats,
                text="0",
                font=("Digital-7", 20),
                bg="black",
                fg="#00ff22",
                width=4,
                height=1)
mindistlabel.place(x=460,y=80)
maxdist=Label(stats,
                text="Maximum Distance:",
                font=("Arial", 15),
                bg="#00ff95")
maxdist.place(x=280,y=130)
maxdistlabel=Label(stats,
                text="0",
                font=("Digital-7", 20),
                bg="black",
                fg="#00ff22",
                width=4,
                height=1)
maxdistlabel.place(x=460,y=130)
avgdist=Label(stats,
                text="Average Distance:",
                font=("Arial", 15),
                bg="#00ff95")
avgdist.place(x=280,y=180)
avgdistlabel=Label(stats,
                text="0",
                font=("Digital-7", 20),
                bg="black",
                fg="#00ff22",
                width=4,
                height=1)
avgdistlabel.place(x=460,y=180)
minspeed=Label(stats,
                text="Minimum Speed:",
                font=("Arial", 15),
                bg="#00ff95")
minspeed.place(x=800,y=80)
minspeedlabel=Label(stats,
                text="0",
                font=("Digital-7", 20),
                bg="black",
                fg="#00ff22",
                width=4,
                height=1)
minspeedlabel.place(x=980,y=80)
maxspeed=Label(stats,
                text="Maximum Speed:",
                font=("Arial", 15),
                bg="#00ff95")
maxspeed.place(x=800,y=130)
maxspeedlabel=Label(stats,
                text="0",
                font=("Digital-7", 20),
                bg="black",
                fg="#00ff22",
                width=4,
                height=1)
maxspeedlabel.place(x=980,y=130)
avgspeed=Label(stats,
                text="Average Speed:",
                font=("Arial", 15),
                bg="#00ff95")
avgspeed.place(x=800,y=180)
avgspeedlabel=Label(stats,
                text="0",
                font=("Digital-7", 20),
                bg="black",
                fg="#00ff22",
                width=4,
                height=1)
avgspeedlabel.place(x=980,y=180)
fullgraph=Button(stats,
                text="Full Graphs",
                font=("Arial",15),
                command=fullgraph)
fullgraph.place(x=220,y=260)
unitlabel=Label(stats,
                text="Time: Minutes | Distance: Kilometers | Speed: Kilometers per Hour")
unitlabel.place(x=710,y=280)
update()
window.mainloop()
